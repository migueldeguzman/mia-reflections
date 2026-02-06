#!/usr/bin/env python3
"""Generate Vesla ERP Cloud Cost Projection spreadsheet with live formulas."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============================================================
# STYLES
# ============================================================
header_font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
section_font = Font(name="Calibri", bold=True, size=11, color="2F5496")
section_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
input_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Yellow = editable
calc_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")   # Green = calculated
total_fill = PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid")   # Orange = totals
bold = Font(name="Calibri", bold=True, size=11)
normal = Font(name="Calibri", size=11)
money_fmt = '#,##0.00'
pct_fmt = '0%'
int_fmt = '#,##0'
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border

def style_section(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = section_font
        cell.fill = section_fill
        cell.border = thin_border

def style_row(ws, row, max_col, fill=None, fmt=None, font=None):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.border = thin_border
        cell.font = font or normal
        if fill:
            cell.fill = fill
        if fmt and c >= 3:
            cell.number_format = fmt

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ============================================================
# SHEET 1: INPUTS & ASSUMPTIONS
# ============================================================
ws1 = wb.active
ws1.title = "Inputs"
set_col_widths(ws1, [40, 20, 18, 18, 50])

# Title
ws1.merge_cells('A1:E1')
ws1['A1'] = "VESLA ERP — Cloud Cost Projection (Neon + Render)"
ws1['A1'].font = Font(name="Calibri", bold=True, size=14, color="2F5496")
ws1['A2'] = "Yellow cells = editable inputs. Change them and all calculations update automatically."
ws1['A2'].font = Font(name="Calibri", italic=True, size=10, color="808080")

# --- Company Assumptions ---
r = 4
ws1.cell(r, 1, "COMPANY ASSUMPTIONS"); style_section(ws1, r, 5)
r += 1
headers = ["Parameter", "Value", "", "", "Notes"]
for c, h in enumerate(headers, 1):
    ws1.cell(r, c, h)
style_header(ws1, r, 5)

inputs = [
    ("Number of companies", 5, None, "Total tenant count"),
    ("Vehicles per company", 200, None, "Medium fleet = 100-300"),
    ("Concurrent users per company (peak)", 20, None, "Staff on system during business hours"),
    ("Active contracts per company/month", 500, None, "Bookings + rentals"),
    ("Invoices per company/month", 500, None, "Billing events"),
    ("TARS fines per company/month", 200, None, "Traffic fines processed"),
    ("Employees per company (HR)", 50, None, "HR/payroll records"),
    ("Customers per company", 2000, None, "Growing customer base"),
]

for i, (label, val, _, note) in enumerate(inputs):
    r += 1
    ws1.cell(r, 1, label).font = normal
    ws1.cell(r, 2, val).font = bold
    ws1.cell(r, 2).fill = input_fill
    ws1.cell(r, 2).number_format = int_fmt
    ws1.cell(r, 5, note).font = Font(name="Calibri", size=10, color="808080")
    style_row(ws1, r, 5)

companies_cell = "B6"   # row 6 = number of companies
r += 1

# --- Usage Pattern ---
r += 1
ws1.cell(r, 1, "USAGE PATTERN"); style_section(ws1, r, 5)
r += 1
headers2 = ["Parameter", "Value", "Unit", "", "Notes"]
for c, h in enumerate(headers2, 1):
    ws1.cell(r, c, h)
style_header(ws1, r, 5)

usage = [
    ("Business hours per day", 14, "hours", "Typical UAE: 8am-10pm"),
    ("Days per month", 30, "days", "Calendar month"),
    ("Off-peak hours per day", 10, "hours", "=24 minus business hours"),
    ("Weekend/holiday reduction", 0.3, "", "30% less traffic on weekends"),
]

usage_start_row = r + 1
for i, (label, val, unit, note) in enumerate(usage):
    r += 1
    ws1.cell(r, 1, label).font = normal
    ws1.cell(r, 2, val).font = bold
    ws1.cell(r, 2).fill = input_fill
    if isinstance(val, float) and val < 1:
        ws1.cell(r, 2).number_format = pct_fmt
    else:
        ws1.cell(r, 2).number_format = int_fmt
    ws1.cell(r, 3, unit).font = normal
    ws1.cell(r, 5, note).font = Font(name="Calibri", size=10, color="808080")
    style_row(ws1, r, 5)

biz_hours_cell = f"B{usage_start_row}"      # 14
days_cell = f"B{usage_start_row + 1}"        # 30
offpeak_hours_cell = f"B{usage_start_row + 2}" # 10
r += 1

# --- Neon Pricing ---
r += 1
ws1.cell(r, 1, "NEON PRICING (from neon.com/docs — Aug 2025 model)"); style_section(ws1, r, 5)
r += 1
for c, h in enumerate(["Parameter", "Launch Plan", "Scale Plan", "", "Source"], 1):
    ws1.cell(r, c, h)
style_header(ws1, r, 5)

neon_prices = [
    ("Compute rate (per CU-hour)", 0.106, 0.222, "neon.com/docs/introduction/plans"),
    ("Storage rate (per GB-month)", 0.35, 0.35, "Same on both plans"),
    ("Instant Restore (per GB-month)", 0.20, 0.20, "For point-in-time recovery"),
    ("Network transfer included", 100, 100, "GB — then $0.10/GB overage"),
    ("Extra branch (per branch-month)", 1.50, 1.50, "Beyond plan included branches"),
]

neon_start_row = r + 1
for i, (label, launch, scale, source) in enumerate(neon_prices):
    r += 1
    ws1.cell(r, 1, label).font = normal
    ws1.cell(r, 2, launch).font = bold
    ws1.cell(r, 2).fill = input_fill
    ws1.cell(r, 3, scale).font = bold
    ws1.cell(r, 3).fill = input_fill
    if isinstance(launch, float):
        ws1.cell(r, 2).number_format = '$#,##0.000'
        ws1.cell(r, 3).number_format = '$#,##0.000'
    else:
        ws1.cell(r, 2).number_format = int_fmt
        ws1.cell(r, 3).number_format = int_fmt
    ws1.cell(r, 5, source).font = Font(name="Calibri", size=10, color="808080")
    style_row(ws1, r, 5)

# Store references
neon_launch_compute = f"B{neon_start_row}"       # $0.106
neon_scale_compute = f"C{neon_start_row}"         # $0.222
neon_storage_rate = f"B{neon_start_row + 1}"      # $0.35
neon_restore_rate = f"B{neon_start_row + 2}"      # $0.20
r += 1

# --- Render Pricing ---
r += 1
ws1.cell(r, 1, "RENDER PRICING (render.com/docs — current tiers)"); style_section(ws1, r, 5)
r += 1
for c, h in enumerate(["Instance Type", "Monthly Cost", "CPU", "RAM", "Notes"], 1):
    ws1.cell(r, c, h)
style_header(ws1, r, 5)

render_tiers = [
    ("Free", 0, "0.1", "512 MB", "Spins down after 15min idle — NOT for production"),
    ("Starter", 7, "0.5", "512 MB", "Hobby / background workers"),
    ("Standard", 25, "1", "2 GB", "Good starting point for 5 medium clients"),
    ("Pro", 85, "2", "4 GB", "When Standard CPU maxes out"),
    ("Pro Plus", 175, "4", "8 GB", "High concurrency / heavy queries"),
    ("Pro Max", 225, "4", "16 GB", "Memory-intensive workloads"),
    ("Pro Ultra", 450, "8", "32 GB", "Enterprise scale"),
]

render_start_row = r + 1
for i, (tier, cost, cpu, ram, note) in enumerate(render_tiers):
    r += 1
    ws1.cell(r, 1, tier).font = normal
    ws1.cell(r, 2, cost).font = bold
    ws1.cell(r, 2).fill = input_fill
    ws1.cell(r, 2).number_format = '$#,##0.00'
    ws1.cell(r, 3, cpu).font = normal
    ws1.cell(r, 4, ram).font = normal
    ws1.cell(r, 5, note).font = Font(name="Calibri", size=10, color="808080")
    style_row(ws1, r, 5)

# Store key render prices
render_starter = f"B{render_start_row + 1}"   # $7
render_standard = f"B{render_start_row + 2}"  # $25
render_pro = f"B{render_start_row + 3}"       # $85

# --- DB Size Estimation ---
r += 2
ws1.cell(r, 1, "DATABASE SIZE ESTIMATION"); style_section(ws1, r, 5)
r += 1
for c, h in enumerate(["Data Category", "GB per Company", "× Companies", "Total GB", "Basis"], 1):
    ws1.cell(r, c, h)
style_header(ws1, r, 5)

db_items = [
    ("Core data (vehicles, contracts, customers, invoices)", 1.8, "275 Prisma models worth of transactional data"),
    ("TARS fines + Salik historical", 0.5, "Monthly fine imports accumulate"),
    ("HR + payroll + attendance", 0.3, "50 employees × daily records"),
    ("Service center + parts inventory", 0.3, "Work orders, parts catalog"),
    ("Index overhead (628 indexes)", 1.5, "~40-50% overhead on data size is typical for heavy indexing"),
]

db_start_row = r + 1
for i, (cat, gb, basis) in enumerate(db_items):
    r += 1
    ws1.cell(r, 1, cat).font = normal
    ws1.cell(r, 2, gb).font = bold
    ws1.cell(r, 2).fill = input_fill
    ws1.cell(r, 2).number_format = '#,##0.0'
    # Col C = companies ref
    ws1.cell(r, 3).font = normal
    ws1.cell(r, 3, f"={companies_cell}")
    ws1.cell(r, 3).number_format = int_fmt
    # Col D = B * C
    ws1.cell(r, 4).font = bold
    ws1.cell(r, 4, f"=B{r}*C{r}")  # formula
    ws1.cell(r, 4).number_format = '#,##0.0'
    ws1.cell(r, 4).fill = calc_fill
    ws1.cell(r, 5, basis).font = Font(name="Calibri", size=10, color="808080")
    style_row(ws1, r, 5)

r += 1
ws1.cell(r, 1, "TOTAL STORAGE").font = bold
ws1.cell(r, 4).font = bold
ws1.cell(r, 4, f"=SUM(D{db_start_row}:D{r-1})")
ws1.cell(r, 4).number_format = '#,##0.0'
ws1.cell(r, 4).fill = total_fill
ws1.cell(r, 5, "GB across all companies").font = Font(name="Calibri", size=10, color="808080")
style_row(ws1, r, 5, font=bold)
total_storage_row = r
total_storage_cell = f"D{r}"


# ============================================================
# SHEET 2: OPTION A — STARTER
# ============================================================
ws2 = wb.create_sheet("Option A — Starter")
set_col_widths(ws2, [45, 18, 18, 18, 40])

ws2.merge_cells('A1:E1')
ws2['A1'] = "OPTION A: Starter Production (~$81/mo)"
ws2['A1'].font = Font(name="Calibri", bold=True, size=14, color="2F5496")
ws2['A2'] = "Best for: First 5 clients onboarding. Neon Launch plan + Render Standard."
ws2['A2'].font = Font(name="Calibri", italic=True, size=10, color="808080")

# --- Neon Compute ---
r = 4
ws2.cell(r, 1, "NEON — LAUNCH PLAN (Compute)"); style_section(ws2, r, 5)
r += 1
for c, h in enumerate(["Component", "Value", "Unit", "Cost", "Formula"], 1):
    ws2.cell(r, c, h)
style_header(ws2, r, 5)

r += 1
ws2.cell(r, 1, "Average CU size (business hours)").font = normal
ws2.cell(r, 2, 1).font = bold; ws2.cell(r, 2).fill = input_fill
ws2.cell(r, 3, "CU").font = normal
a_biz_cu = f"B{r}"
style_row(ws2, r, 5)
a_biz_cu_row = r

r += 1
ws2.cell(r, 1, "Average CU size (off-peak)").font = normal
ws2.cell(r, 2, 0.25).font = bold; ws2.cell(r, 2).fill = input_fill
ws2.cell(r, 3, "CU").font = normal
a_offpeak_cu = f"B{r}"
style_row(ws2, r, 5)
a_offpeak_cu_row = r

r += 1
ws2.cell(r, 1, "Business hours/month").font = normal
ws2.cell(r, 2, f"=Inputs!{biz_hours_cell}*Inputs!{days_cell}")
ws2.cell(r, 2).number_format = int_fmt
ws2.cell(r, 2).fill = calc_fill
ws2.cell(r, 3, "hours").font = normal
ws2.cell(r, 5, f"=Inputs!{biz_hours_cell} hrs × Inputs!{days_cell} days").font = Font(name="Calibri", size=10, color="808080")
style_row(ws2, r, 5)
biz_hrs_mo = f"B{r}"

r += 1
ws2.cell(r, 1, "Off-peak hours/month").font = normal
ws2.cell(r, 2, f"=Inputs!{offpeak_hours_cell}*Inputs!{days_cell}")
ws2.cell(r, 2).number_format = int_fmt
ws2.cell(r, 2).fill = calc_fill
ws2.cell(r, 3, "hours").font = normal
style_row(ws2, r, 5)
offpeak_hrs_mo = f"B{r}"

r += 1
ws2.cell(r, 1, "Business CU-hours").font = normal
ws2.cell(r, 2, f"={a_biz_cu}*{biz_hrs_mo}")
ws2.cell(r, 2).number_format = int_fmt
ws2.cell(r, 2).fill = calc_fill
ws2.cell(r, 3, "CU-hrs").font = normal
ws2.cell(r, 4, f"=B{r}*Inputs!{neon_launch_compute}")
ws2.cell(r, 4).number_format = money_fmt
ws2.cell(r, 4).fill = calc_fill
ws2.cell(r, 5, f"CU-hrs × ${neon_launch_compute} rate").font = Font(name="Calibri", size=10, color="808080")
style_row(ws2, r, 5)
a_biz_cost = f"D{r}"

r += 1
ws2.cell(r, 1, "Off-peak CU-hours").font = normal
ws2.cell(r, 2, f"={a_offpeak_cu}*{offpeak_hrs_mo}")
ws2.cell(r, 2).number_format = int_fmt
ws2.cell(r, 2).fill = calc_fill
ws2.cell(r, 3, "CU-hrs").font = normal
ws2.cell(r, 4, f"=B{r}*Inputs!{neon_launch_compute}")
ws2.cell(r, 4).number_format = money_fmt
ws2.cell(r, 4).fill = calc_fill
style_row(ws2, r, 5)
a_offpeak_cost = f"D{r}"

r += 1
ws2.cell(r, 1, "COMPUTE SUBTOTAL").font = bold
ws2.cell(r, 4, f"={a_biz_cost}+{a_offpeak_cost}")
ws2.cell(r, 4).number_format = money_fmt
ws2.cell(r, 4).fill = total_fill
ws2.cell(r, 4).font = bold
style_row(ws2, r, 5, font=bold)
a_compute_total = f"D{r}"

# --- Neon Storage ---
r += 2
ws2.cell(r, 1, "NEON — LAUNCH PLAN (Storage & Other)"); style_section(ws2, r, 5)
r += 1
for c, h in enumerate(["Component", "Value", "Unit", "Cost", "Formula"], 1):
    ws2.cell(r, c, h)
style_header(ws2, r, 5)

r += 1
ws2.cell(r, 1, "Total storage").font = normal
ws2.cell(r, 2, f"=Inputs!{total_storage_cell}")
ws2.cell(r, 2).number_format = '#,##0.0'
ws2.cell(r, 2).fill = calc_fill
ws2.cell(r, 3, "GB").font = normal
ws2.cell(r, 4, f"=B{r}*Inputs!{neon_storage_rate}")
ws2.cell(r, 4).number_format = money_fmt
ws2.cell(r, 4).fill = calc_fill
style_row(ws2, r, 5)
a_storage_cost = f"D{r}"

r += 1
ws2.cell(r, 1, "Instant Restore (est. monthly changes)").font = normal
ws2.cell(r, 2, 5).font = bold; ws2.cell(r, 2).fill = input_fill
ws2.cell(r, 2).number_format = '#,##0.0'
ws2.cell(r, 3, "GB").font = normal
ws2.cell(r, 4, f"=B{r}*Inputs!{neon_restore_rate}")
ws2.cell(r, 4).number_format = money_fmt
ws2.cell(r, 4).fill = calc_fill
style_row(ws2, r, 5)
a_restore_cost = f"D{r}"

r += 1
ws2.cell(r, 1, "Network transfer").font = normal
ws2.cell(r, 2, "< 100 GB included").font = normal
ws2.cell(r, 4, 0).number_format = money_fmt
ws2.cell(r, 4).fill = calc_fill
style_row(ws2, r, 5)
a_network_cost = f"D{r}"

r += 1
ws2.cell(r, 1, "NEON TOTAL").font = bold
ws2.cell(r, 4, f"={a_compute_total}+{a_storage_cost}+{a_restore_cost}+{a_network_cost}")
ws2.cell(r, 4).number_format = money_fmt
ws2.cell(r, 4).fill = total_fill
ws2.cell(r, 4).font = bold
style_row(ws2, r, 5, font=bold)
a_neon_total = f"D{r}"

# --- Render ---
r += 2
ws2.cell(r, 1, "RENDER — SERVICES"); style_section(ws2, r, 5)
r += 1
for c, h in enumerate(["Service", "Instance Type", "Instances", "Cost", "Notes"], 1):
    ws2.cell(r, c, h)
style_header(ws2, r, 5)

r += 1
ws2.cell(r, 1, "Backend API (Node + Express + Prisma)").font = normal
ws2.cell(r, 2, "Standard").font = normal
ws2.cell(r, 3, 1).font = normal; ws2.cell(r, 3).fill = input_fill
ws2.cell(r, 4, f"=C{r}*Inputs!{render_standard}")
ws2.cell(r, 4).number_format = money_fmt; ws2.cell(r, 4).fill = calc_fill
ws2.cell(r, 5, "1 CPU, 2 GB RAM").font = Font(name="Calibri", size=10, color="808080")
style_row(ws2, r, 5)
a_render_backend = f"D{r}"

r += 1
ws2.cell(r, 1, "Frontend (React + Vite)").font = normal
ws2.cell(r, 2, "Static Site").font = normal
ws2.cell(r, 3, 1).font = normal
ws2.cell(r, 4, 0)
ws2.cell(r, 4).number_format = money_fmt; ws2.cell(r, 4).fill = calc_fill
ws2.cell(r, 5, "Free on Render").font = Font(name="Calibri", size=10, color="808080")
style_row(ws2, r, 5)
a_render_frontend = f"D{r}"

r += 1
ws2.cell(r, 1, "Background Worker (TARS, Speed Sync, crons)").font = normal
ws2.cell(r, 2, "Starter").font = normal
ws2.cell(r, 3, 1).font = normal; ws2.cell(r, 3).fill = input_fill
ws2.cell(r, 4, f"=C{r}*Inputs!{render_starter}")
ws2.cell(r, 4).number_format = money_fmt; ws2.cell(r, 4).fill = calc_fill
ws2.cell(r, 5, "0.5 CPU, 512 MB").font = Font(name="Calibri", size=10, color="808080")
style_row(ws2, r, 5)
a_render_worker = f"D{r}"

r += 1
ws2.cell(r, 1, "RENDER TOTAL").font = bold
ws2.cell(r, 4, f"={a_render_backend}+{a_render_frontend}+{a_render_worker}")
ws2.cell(r, 4).number_format = money_fmt
ws2.cell(r, 4).fill = total_fill
ws2.cell(r, 4).font = bold
style_row(ws2, r, 5, font=bold)
a_render_total = f"D{r}"

# --- Grand Total ---
r += 2
ws2.cell(r, 1, "OPTION A — GRAND TOTAL").font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
ws2.cell(r, 1).fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
for c in range(1, 6):
    ws2.cell(r, c).fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    ws2.cell(r, c).border = thin_border

r += 1
ws2.cell(r, 1, "Neon (Database)").font = bold
ws2.cell(r, 4, f"={a_neon_total}").font = bold
ws2.cell(r, 4).number_format = money_fmt
style_row(ws2, r, 5)

r += 1
ws2.cell(r, 1, "Render (Hosting)").font = bold
ws2.cell(r, 4, f"={a_render_total}").font = bold
ws2.cell(r, 4).number_format = money_fmt
style_row(ws2, r, 5)

r += 1
ws2.cell(r, 1, "TOTAL PER MONTH").font = Font(name="Calibri", bold=True, size=12)
ws2.cell(r, 4, f"={a_neon_total}+{a_render_total}")
ws2.cell(r, 4).number_format = '$#,##0.00'
ws2.cell(r, 4).fill = total_fill
ws2.cell(r, 4).font = Font(name="Calibri", bold=True, size=12)
style_row(ws2, r, 5, font=Font(name="Calibri", bold=True, size=12))
a_grand = f"D{r}"

r += 1
ws2.cell(r, 1, "TOTAL PER YEAR").font = bold
ws2.cell(r, 4, f"={a_grand}*12")
ws2.cell(r, 4).number_format = '$#,##0.00'
ws2.cell(r, 4).fill = total_fill
style_row(ws2, r, 5, font=bold)

r += 1
ws2.cell(r, 1, "PER COMPANY / MONTH").font = bold
ws2.cell(r, 4, f"={a_grand}/Inputs!{companies_cell}")
ws2.cell(r, 4).number_format = '$#,##0.00'
ws2.cell(r, 4).fill = total_fill
style_row(ws2, r, 5, font=bold)


# ============================================================
# SHEET 3: OPTION B — GROWTH
# ============================================================
ws3 = wb.create_sheet("Option B — Growth")
set_col_widths(ws3, [45, 18, 18, 18, 40])

ws3.merge_cells('A1:E1')
ws3['A1'] = "OPTION B: Growth Production (~$244/mo)"
ws3['A1'].font = Font(name="Calibri", bold=True, size=14, color="2F5496")
ws3['A2'] = "Best for: 5 active companies using daily. Neon Launch (higher CU) + Render Pro."
ws3['A2'].font = Font(name="Calibri", italic=True, size=10, color="808080")

r = 4
ws3.cell(r, 1, "NEON — LAUNCH PLAN (Higher Compute)"); style_section(ws3, r, 5)
r += 1
for c, h in enumerate(["Component", "Value", "Unit", "Cost", "Formula"], 1):
    ws3.cell(r, c, h)
style_header(ws3, r, 5)

r += 1
ws3.cell(r, 1, "Average CU size (business hours)").font = normal
ws3.cell(r, 2, 2).font = bold; ws3.cell(r, 2).fill = input_fill
ws3.cell(r, 3, "CU").font = normal
b_biz_cu = f"B{r}"
style_row(ws3, r, 5)

r += 1
ws3.cell(r, 1, "Average CU size (off-peak)").font = normal
ws3.cell(r, 2, 0.5).font = bold; ws3.cell(r, 2).fill = input_fill
ws3.cell(r, 3, "CU").font = normal
b_offpeak_cu = f"B{r}"
style_row(ws3, r, 5)

r += 1
ws3.cell(r, 1, "Business hours/month").font = normal
ws3.cell(r, 2, f"=Inputs!{biz_hours_cell}*Inputs!{days_cell}")
ws3.cell(r, 2).number_format = int_fmt; ws3.cell(r, 2).fill = calc_fill
ws3.cell(r, 3, "hours").font = normal
style_row(ws3, r, 5)
b_biz_hrs = f"B{r}"

r += 1
ws3.cell(r, 1, "Off-peak hours/month").font = normal
ws3.cell(r, 2, f"=Inputs!{offpeak_hours_cell}*Inputs!{days_cell}")
ws3.cell(r, 2).number_format = int_fmt; ws3.cell(r, 2).fill = calc_fill
ws3.cell(r, 3, "hours").font = normal
style_row(ws3, r, 5)
b_offpeak_hrs = f"B{r}"

r += 1
ws3.cell(r, 1, "Business CU-hours").font = normal
ws3.cell(r, 2, f"={b_biz_cu}*{b_biz_hrs}")
ws3.cell(r, 2).number_format = int_fmt; ws3.cell(r, 2).fill = calc_fill
ws3.cell(r, 3, "CU-hrs").font = normal
ws3.cell(r, 4, f"=B{r}*Inputs!{neon_launch_compute}")
ws3.cell(r, 4).number_format = money_fmt; ws3.cell(r, 4).fill = calc_fill
style_row(ws3, r, 5)
b_biz_cost = f"D{r}"

r += 1
ws3.cell(r, 1, "Off-peak CU-hours").font = normal
ws3.cell(r, 2, f"={b_offpeak_cu}*{b_offpeak_hrs}")
ws3.cell(r, 2).number_format = int_fmt; ws3.cell(r, 2).fill = calc_fill
ws3.cell(r, 3, "CU-hrs").font = normal
ws3.cell(r, 4, f"=B{r}*Inputs!{neon_launch_compute}")
ws3.cell(r, 4).number_format = money_fmt; ws3.cell(r, 4).fill = calc_fill
style_row(ws3, r, 5)
b_offpeak_cost = f"D{r}"

r += 1
ws3.cell(r, 1, "COMPUTE SUBTOTAL").font = bold
ws3.cell(r, 4, f"={b_biz_cost}+{b_offpeak_cost}")
ws3.cell(r, 4).number_format = money_fmt; ws3.cell(r, 4).fill = total_fill; ws3.cell(r, 4).font = bold
style_row(ws3, r, 5, font=bold)
b_compute_total = f"D{r}"

r += 2
ws3.cell(r, 1, "NEON — Storage & Other"); style_section(ws3, r, 5)
r += 1
for c, h in enumerate(["Component", "Value", "Unit", "Cost", ""], 1):
    ws3.cell(r, c, h)
style_header(ws3, r, 5)

r += 1
ws3.cell(r, 1, "Total storage").font = normal
ws3.cell(r, 2, f"=Inputs!{total_storage_cell}")
ws3.cell(r, 2).number_format = '#,##0.0'; ws3.cell(r, 2).fill = calc_fill
ws3.cell(r, 3, "GB").font = normal
ws3.cell(r, 4, f"=B{r}*Inputs!{neon_storage_rate}")
ws3.cell(r, 4).number_format = money_fmt; ws3.cell(r, 4).fill = calc_fill
style_row(ws3, r, 5)
b_storage = f"D{r}"

r += 1
ws3.cell(r, 1, "Instant Restore (est. changes)").font = normal
ws3.cell(r, 2, 8).font = bold; ws3.cell(r, 2).fill = input_fill
ws3.cell(r, 3, "GB").font = normal
ws3.cell(r, 4, f"=B{r}*Inputs!{neon_restore_rate}")
ws3.cell(r, 4).number_format = money_fmt; ws3.cell(r, 4).fill = calc_fill
style_row(ws3, r, 5)
b_restore = f"D{r}"

r += 1
ws3.cell(r, 1, "NEON TOTAL").font = bold
ws3.cell(r, 4, f"={b_compute_total}+{b_storage}+{b_restore}")
ws3.cell(r, 4).number_format = money_fmt; ws3.cell(r, 4).fill = total_fill; ws3.cell(r, 4).font = bold
style_row(ws3, r, 5, font=bold)
b_neon_total = f"D{r}"

# Render for Option B
r += 2
ws3.cell(r, 1, "RENDER — SERVICES"); style_section(ws3, r, 5)
r += 1
for c, h in enumerate(["Service", "Instance Type", "Instances", "Cost", "Notes"], 1):
    ws3.cell(r, c, h)
style_header(ws3, r, 5)

r += 1
ws3.cell(r, 1, "Backend API").font = normal
ws3.cell(r, 2, "Pro").font = normal
ws3.cell(r, 3, 1); ws3.cell(r, 3).fill = input_fill
ws3.cell(r, 4, f"=C{r}*Inputs!{render_pro}")
ws3.cell(r, 4).number_format = money_fmt; ws3.cell(r, 4).fill = calc_fill
style_row(ws3, r, 5)
b_r_backend = f"D{r}"

r += 1
ws3.cell(r, 1, "Frontend (Static Site)").font = normal
ws3.cell(r, 2, "Free").font = normal
ws3.cell(r, 4, 0); ws3.cell(r, 4).number_format = money_fmt; ws3.cell(r, 4).fill = calc_fill
style_row(ws3, r, 5)
b_r_frontend = f"D{r}"

r += 1
ws3.cell(r, 1, "Background Worker").font = normal
ws3.cell(r, 2, "Standard").font = normal
ws3.cell(r, 3, 1); ws3.cell(r, 3).fill = input_fill
ws3.cell(r, 4, f"=C{r}*Inputs!{render_standard}")
ws3.cell(r, 4).number_format = money_fmt; ws3.cell(r, 4).fill = calc_fill
style_row(ws3, r, 5)
b_r_worker = f"D{r}"

r += 1
ws3.cell(r, 1, "Team workspace ($19/seat)").font = normal
ws3.cell(r, 2, "Team").font = normal
ws3.cell(r, 3, 1); ws3.cell(r, 3).fill = input_fill
ws3.cell(r, 4, f"=C{r}*19")
ws3.cell(r, 4).number_format = money_fmt; ws3.cell(r, 4).fill = calc_fill
style_row(ws3, r, 5)
b_r_team = f"D{r}"

r += 1
ws3.cell(r, 1, "RENDER TOTAL").font = bold
ws3.cell(r, 4, f"={b_r_backend}+{b_r_frontend}+{b_r_worker}+{b_r_team}")
ws3.cell(r, 4).number_format = money_fmt; ws3.cell(r, 4).fill = total_fill; ws3.cell(r, 4).font = bold
style_row(ws3, r, 5, font=bold)
b_render_total = f"D{r}"

# Grand Total B
r += 2
ws3.cell(r, 1, "OPTION B — GRAND TOTAL").font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
for c in range(1, 6):
    ws3.cell(r, c).fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    ws3.cell(r, c).border = thin_border

r += 1
ws3.cell(r, 1, "Neon (Database)").font = bold
ws3.cell(r, 4, f"={b_neon_total}").font = bold; ws3.cell(r, 4).number_format = money_fmt
style_row(ws3, r, 5)
r += 1
ws3.cell(r, 1, "Render (Hosting)").font = bold
ws3.cell(r, 4, f"={b_render_total}").font = bold; ws3.cell(r, 4).number_format = money_fmt
style_row(ws3, r, 5)
r += 1
ws3.cell(r, 1, "TOTAL PER MONTH").font = Font(name="Calibri", bold=True, size=12)
ws3.cell(r, 4, f"={b_neon_total}+{b_render_total}")
ws3.cell(r, 4).number_format = '$#,##0.00'; ws3.cell(r, 4).fill = total_fill
ws3.cell(r, 4).font = Font(name="Calibri", bold=True, size=12)
style_row(ws3, r, 5, font=Font(name="Calibri", bold=True, size=12))
b_grand = f"D{r}"
r += 1
ws3.cell(r, 1, "TOTAL PER YEAR").font = bold
ws3.cell(r, 4, f"={b_grand}*12"); ws3.cell(r, 4).number_format = '$#,##0.00'; ws3.cell(r, 4).fill = total_fill
style_row(ws3, r, 5, font=bold)
r += 1
ws3.cell(r, 1, "PER COMPANY / MONTH").font = bold
ws3.cell(r, 4, f"={b_grand}/Inputs!{companies_cell}")
ws3.cell(r, 4).number_format = '$#,##0.00'; ws3.cell(r, 4).fill = total_fill
style_row(ws3, r, 5, font=bold)


# ============================================================
# SHEET 4: OPTION C — ENTERPRISE
# ============================================================
ws4 = wb.create_sheet("Option C — Enterprise")
set_col_widths(ws4, [45, 18, 18, 18, 40])

ws4.merge_cells('A1:E1')
ws4['A1'] = "OPTION C: Enterprise / SLA (~$671/mo)"
ws4['A1'].font = Font(name="Calibri", bold=True, size=14, color="2F5496")
ws4['A2'] = "Best for: SLA commitments, compliance. Neon Scale plan (always-on) + Render Pro (×2)."
ws4['A2'].font = Font(name="Calibri", italic=True, size=10, color="808080")

r = 4
ws4.cell(r, 1, "NEON — SCALE PLAN (Always-On Compute)"); style_section(ws4, r, 5)
r += 1
for c, h in enumerate(["Component", "Value", "Unit", "Cost", "Formula"], 1):
    ws4.cell(r, c, h)
style_header(ws4, r, 5)

r += 1
ws4.cell(r, 1, "Primary compute (always-on)").font = normal
ws4.cell(r, 2, 2).font = bold; ws4.cell(r, 2).fill = input_fill
ws4.cell(r, 3, "CU").font = normal
c_cu = f"B{r}"
style_row(ws4, r, 5)

r += 1
ws4.cell(r, 1, "Hours/month (always-on)").font = normal
ws4.cell(r, 2, 730)
ws4.cell(r, 2).number_format = int_fmt; ws4.cell(r, 2).fill = calc_fill
ws4.cell(r, 3, "hours").font = normal
ws4.cell(r, 5, "24 × 30.4 ≈ 730").font = Font(name="Calibri", size=10, color="808080")
style_row(ws4, r, 5)
c_hrs = f"B{r}"

r += 1
ws4.cell(r, 1, "Primary CU-hours").font = normal
ws4.cell(r, 2, f"={c_cu}*{c_hrs}")
ws4.cell(r, 2).number_format = int_fmt; ws4.cell(r, 2).fill = calc_fill
ws4.cell(r, 3, "CU-hrs").font = normal
ws4.cell(r, 4, f"=B{r}*Inputs!{neon_scale_compute}")
ws4.cell(r, 4).number_format = money_fmt; ws4.cell(r, 4).fill = calc_fill
style_row(ws4, r, 5)
c_primary_cost = f"D{r}"

r += 1
ws4.cell(r, 1, "Read replica (reporting queries)").font = normal
ws4.cell(r, 2, 0.5).font = bold; ws4.cell(r, 2).fill = input_fill
ws4.cell(r, 3, "CU").font = normal
c_replica_cu = f"B{r}"
style_row(ws4, r, 5)

r += 1
ws4.cell(r, 1, "Replica CU-hours").font = normal
ws4.cell(r, 2, f"={c_replica_cu}*{c_hrs}")
ws4.cell(r, 2).number_format = int_fmt; ws4.cell(r, 2).fill = calc_fill
ws4.cell(r, 3, "CU-hrs").font = normal
ws4.cell(r, 4, f"=B{r}*Inputs!{neon_scale_compute}")
ws4.cell(r, 4).number_format = money_fmt; ws4.cell(r, 4).fill = calc_fill
style_row(ws4, r, 5)
c_replica_cost = f"D{r}"

r += 1
ws4.cell(r, 1, "COMPUTE SUBTOTAL").font = bold
ws4.cell(r, 4, f"={c_primary_cost}+{c_replica_cost}")
ws4.cell(r, 4).number_format = money_fmt; ws4.cell(r, 4).fill = total_fill; ws4.cell(r, 4).font = bold
style_row(ws4, r, 5, font=bold)
c_compute_total = f"D{r}"

r += 2
ws4.cell(r, 1, "NEON — Storage & Other"); style_section(ws4, r, 5)
r += 1
for c2, h in enumerate(["Component", "Value", "Unit", "Cost", ""], 1):
    ws4.cell(r, c2, h)
style_header(ws4, r, 5)

r += 1
ws4.cell(r, 1, "Storage").font = normal
ws4.cell(r, 2, f"=Inputs!{total_storage_cell}")
ws4.cell(r, 2).number_format = '#,##0.0'; ws4.cell(r, 2).fill = calc_fill
ws4.cell(r, 3, "GB").font = normal
ws4.cell(r, 4, f"=B{r}*Inputs!{neon_storage_rate}")
ws4.cell(r, 4).number_format = money_fmt; ws4.cell(r, 4).fill = calc_fill
style_row(ws4, r, 5)
c_storage = f"D{r}"

r += 1
ws4.cell(r, 1, "Instant Restore (30-day window)").font = normal
ws4.cell(r, 2, 25).font = bold; ws4.cell(r, 2).fill = input_fill
ws4.cell(r, 3, "GB").font = normal
ws4.cell(r, 4, f"=B{r}*Inputs!{neon_restore_rate}")
ws4.cell(r, 4).number_format = money_fmt; ws4.cell(r, 4).fill = calc_fill
style_row(ws4, r, 5)
c_restore = f"D{r}"

r += 1
ws4.cell(r, 1, "NEON TOTAL").font = bold
ws4.cell(r, 4, f"={c_compute_total}+{c_storage}+{c_restore}")
ws4.cell(r, 4).number_format = money_fmt; ws4.cell(r, 4).fill = total_fill; ws4.cell(r, 4).font = bold
style_row(ws4, r, 5, font=bold)
c_neon_total = f"D{r}"

# Render for Option C
r += 2
ws4.cell(r, 1, "RENDER — SERVICES"); style_section(ws4, r, 5)
r += 1
for c2, h in enumerate(["Service", "Instance Type", "Instances", "Cost", "Notes"], 1):
    ws4.cell(r, c2, h)
style_header(ws4, r, 5)

r += 1
ws4.cell(r, 1, "Backend API (load balanced)").font = normal
ws4.cell(r, 2, "Pro").font = normal
ws4.cell(r, 3, 2).font = bold; ws4.cell(r, 3).fill = input_fill
ws4.cell(r, 4, f"=C{r}*Inputs!{render_pro}")
ws4.cell(r, 4).number_format = money_fmt; ws4.cell(r, 4).fill = calc_fill
style_row(ws4, r, 5)
c_r_backend = f"D{r}"

r += 1
ws4.cell(r, 1, "Frontend (Static Site)").font = normal
ws4.cell(r, 2, "Free").font = normal
ws4.cell(r, 4, 0); ws4.cell(r, 4).number_format = money_fmt; ws4.cell(r, 4).fill = calc_fill
style_row(ws4, r, 5)
c_r_frontend = f"D{r}"

r += 1
ws4.cell(r, 1, "Background Worker").font = normal
ws4.cell(r, 2, "Standard").font = normal
ws4.cell(r, 3, 1); ws4.cell(r, 3).fill = input_fill
ws4.cell(r, 4, f"=C{r}*Inputs!{render_standard}")
ws4.cell(r, 4).number_format = money_fmt; ws4.cell(r, 4).fill = calc_fill
style_row(ws4, r, 5)
c_r_worker = f"D{r}"

r += 1
ws4.cell(r, 1, "Team workspace ($19/seat × 3)").font = normal
ws4.cell(r, 2, "Team").font = normal
ws4.cell(r, 3, 3).font = bold; ws4.cell(r, 3).fill = input_fill
ws4.cell(r, 4, f"=C{r}*19")
ws4.cell(r, 4).number_format = money_fmt; ws4.cell(r, 4).fill = calc_fill
style_row(ws4, r, 5)
c_r_team = f"D{r}"

r += 1
ws4.cell(r, 1, "RENDER TOTAL").font = bold
ws4.cell(r, 4, f"={c_r_backend}+{c_r_frontend}+{c_r_worker}+{c_r_team}")
ws4.cell(r, 4).number_format = money_fmt; ws4.cell(r, 4).fill = total_fill; ws4.cell(r, 4).font = bold
style_row(ws4, r, 5, font=bold)
c_render_total = f"D{r}"

# Grand Total C
r += 2
ws4.cell(r, 1, "OPTION C — GRAND TOTAL").font = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
for c2 in range(1, 6):
    ws4.cell(r, c2).fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    ws4.cell(r, c2).border = thin_border

r += 1
ws4.cell(r, 1, "Neon (Database)").font = bold
ws4.cell(r, 4, f"={c_neon_total}").font = bold; ws4.cell(r, 4).number_format = money_fmt
style_row(ws4, r, 5)
r += 1
ws4.cell(r, 1, "Render (Hosting)").font = bold
ws4.cell(r, 4, f"={c_render_total}").font = bold; ws4.cell(r, 4).number_format = money_fmt
style_row(ws4, r, 5)
r += 1
ws4.cell(r, 1, "TOTAL PER MONTH").font = Font(name="Calibri", bold=True, size=12)
ws4.cell(r, 4, f"={c_neon_total}+{c_render_total}")
ws4.cell(r, 4).number_format = '$#,##0.00'; ws4.cell(r, 4).fill = total_fill
ws4.cell(r, 4).font = Font(name="Calibri", bold=True, size=12)
style_row(ws4, r, 5, font=Font(name="Calibri", bold=True, size=12))
c_grand = f"D{r}"
r += 1
ws4.cell(r, 1, "TOTAL PER YEAR").font = bold
ws4.cell(r, 4, f"={c_grand}*12"); ws4.cell(r, 4).number_format = '$#,##0.00'; ws4.cell(r, 4).fill = total_fill
style_row(ws4, r, 5, font=bold)
r += 1
ws4.cell(r, 1, "PER COMPANY / MONTH").font = bold
ws4.cell(r, 4, f"={c_grand}/Inputs!{companies_cell}")
ws4.cell(r, 4).number_format = '$#,##0.00'; ws4.cell(r, 4).fill = total_fill
style_row(ws4, r, 5, font=bold)


# ============================================================
# SHEET 5: COMPARISON SUMMARY
# ============================================================
ws5 = wb.create_sheet("Summary")
set_col_widths(ws5, [30, 20, 20, 20])

ws5.merge_cells('A1:D1')
ws5['A1'] = "SIDE-BY-SIDE COMPARISON"
ws5['A1'].font = Font(name="Calibri", bold=True, size=14, color="2F5496")

r = 3
for c, h in enumerate(["Metric", "Option A (Starter)", "Option B (Growth)", "Option C (Enterprise)"], 1):
    ws5.cell(r, c, h)
style_header(ws5, r, 4)

# References to grand totals on each sheet
# Option A grand total
a_ref = f"'Option A — Starter'!{a_grand}"
b_ref = f"'Option B — Growth'!{b_grand}"
c_ref = f"'Option C — Enterprise'!{c_grand}"

a_neon_ref = f"'Option A — Starter'!{a_neon_total}"
b_neon_ref = f"'Option B — Growth'!{b_neon_total}"
c_neon_ref = f"'Option C — Enterprise'!{c_neon_total}"

a_render_ref = f"'Option A — Starter'!{a_render_total}"
b_render_ref = f"'Option B — Growth'!{b_render_total}"
c_render_ref = f"'Option C — Enterprise'!{c_render_total}"

rows_data = [
    ("Neon (Database) /mo", a_neon_ref, b_neon_ref, c_neon_ref),
    ("Render (Hosting) /mo", a_render_ref, b_render_ref, c_render_ref),
    ("TOTAL / MONTH", a_ref, b_ref, c_ref),
    ("TOTAL / YEAR", f"={a_ref}*12", f"={b_ref}*12", f"={c_ref}*12"),
    ("Per Company / Month", f"={a_ref}/Inputs!{companies_cell}", f"={b_ref}/Inputs!{companies_cell}", f"={c_ref}/Inputs!{companies_cell}"),
    ("Per Company / Year", f"={a_ref}*12/Inputs!{companies_cell}", f"={b_ref}*12/Inputs!{companies_cell}", f"={c_ref}*12/Inputs!{companies_cell}"),
]

for label, va, vb, vc in rows_data:
    r += 1
    ws5.cell(r, 1, label).font = bold if "TOTAL" in label else normal
    for col_idx, val in enumerate([va, vb, vc], 2):
        cell = ws5.cell(r, col_idx)
        if val.startswith("="):
            cell.value = val
        else:
            cell.value = f"={val}"
        cell.number_format = '$#,##0.00'
        cell.alignment = Alignment(horizontal='center')
        if "TOTAL / MONTH" in label:
            cell.fill = total_fill
            cell.font = Font(name="Calibri", bold=True, size=12)
        elif "TOTAL" in label:
            cell.fill = total_fill
            cell.font = bold
    style_row(ws5, r, 4)

# Add descriptive rows
r += 2
ws5.cell(r, 1, "Best for:").font = section_font
ws5.cell(r, 2, "Onboarding / MVP").font = normal; ws5.cell(r, 2).alignment = Alignment(horizontal='center')
ws5.cell(r, 3, "Active daily use").font = normal; ws5.cell(r, 3).alignment = Alignment(horizontal='center')
ws5.cell(r, 4, "SLA / compliance").font = normal; ws5.cell(r, 4).alignment = Alignment(horizontal='center')

r += 1
ws5.cell(r, 1, "Backend tier:").font = section_font
ws5.cell(r, 2, "Standard (1 CPU, 2GB)").font = normal; ws5.cell(r, 2).alignment = Alignment(horizontal='center')
ws5.cell(r, 3, "Pro (2 CPU, 4GB)").font = normal; ws5.cell(r, 3).alignment = Alignment(horizontal='center')
ws5.cell(r, 4, "Pro ×2 (load balanced)").font = normal; ws5.cell(r, 4).alignment = Alignment(horizontal='center')

r += 1
ws5.cell(r, 1, "DB compute:").font = section_font
ws5.cell(r, 2, "1 CU (autoscale)").font = normal; ws5.cell(r, 2).alignment = Alignment(horizontal='center')
ws5.cell(r, 3, "2 CU (autoscale)").font = normal; ws5.cell(r, 3).alignment = Alignment(horizontal='center')
ws5.cell(r, 4, "2 CU + replica (always-on)").font = normal; ws5.cell(r, 4).alignment = Alignment(horizontal='center')

r += 1
ws5.cell(r, 1, "Neon plan:").font = section_font
ws5.cell(r, 2, "Launch").font = normal; ws5.cell(r, 2).alignment = Alignment(horizontal='center')
ws5.cell(r, 3, "Launch").font = normal; ws5.cell(r, 3).alignment = Alignment(horizontal='center')
ws5.cell(r, 4, "Scale (SOC2, SLA)").font = normal; ws5.cell(r, 4).alignment = Alignment(horizontal='center')

r += 2
ws5.cell(r, 1, "💡 Change any yellow cell on the Inputs sheet and all numbers recalculate.").font = Font(name="Calibri", italic=True, size=11, color="2F5496")
ws5.merge_cells(f'A{r}:D{r}')


# ============================================================
# SAVE
# ============================================================
out_path = "/Users/miguelitodeguzman/Projects/tech-project/docs/Vesla_ERP_Cloud_Cost_Projection.xlsx"
wb.save(out_path)
print(f"✅ Saved to {out_path}")
