#!/usr/bin/env python3
"""
Vesla ERP Cloud Cost Projection v2
Quantity × Unit Price model — no lump sums.
Everything flows: records → row size → GB → GB × rate = cost
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# Styles
hdr_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
hdr_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
sec_font = Font(name="Calibri", bold=True, size=11, color="2F5496")
sec_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
inp_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")   # Yellow = input
calc_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Green = calc
tot_fill = PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid")   # Orange = total
bold = Font(name="Calibri", bold=True, size=11)
norm = Font(name="Calibri", size=11)
note_font = Font(name="Calibri", size=9, color="808080")
bdr = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
money = '#,##0.00'
money3 = '$#,##0.000'
int_f = '#,##0'
dec1 = '#,##0.0'
dec3 = '#,##0.000'
pct = '0%'

def hdr(ws, r, cols):
    for c in range(1, cols+1):
        cl = ws.cell(r, c)
        cl.font = hdr_font; cl.fill = hdr_fill; cl.border = bdr
        cl.alignment = Alignment(horizontal='center', wrap_text=True)

def sec(ws, r, cols, text):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols)
    ws.cell(r, 1, text).font = sec_font
    for c in range(1, cols+1):
        ws.cell(r, c).fill = sec_fill; ws.cell(r, c).border = bdr

def cel(ws, r, c, val=None, font=None, fill=None, fmt=None, align=None):
    cl = ws.cell(r, c)
    if val is not None: cl.value = val
    cl.font = font or norm; cl.border = bdr
    if fill: cl.fill = fill
    if fmt: cl.number_format = fmt
    if align: cl.alignment = Alignment(horizontal=align)
    return cl

def widths(ws, w_list):
    for i, w in enumerate(w_list, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ================================================================
# SHEET 1: UNIT PRICES
# ================================================================
ws = wb.active
ws.title = "Unit Prices"
widths(ws, [35, 18, 14, 14, 35])
MC = 5

ws.merge_cells('A1:E1')
cel(ws, 1, 1, "UNIT PRICES — Source of Truth", Font(name="Calibri", bold=True, size=14, color="2F5496"))
cel(ws, 2, 1, "All yellow cells are editable. Every cost in this workbook = Quantity × Unit Price from this sheet.", note_font)

# Neon rates
r = 4; sec(ws, r, MC, "NEON RATES (neon.com/docs/introduction/plans — Aug 2025 pricing)")
r += 1
for c, h in enumerate(["Item", "Launch Plan", "Scale Plan", "", "Source"], 1):
    cel(ws, r, c, h)
hdr(ws, r, MC)

neon = [
    ("Compute (per CU-hour)", 0.106, 0.222, "CU = ~4 GB RAM + CPU + SSD"),
    ("Storage (per GB-month)", 0.35, 0.35, "Actual data size, metered hourly"),
    ("Instant Restore (per GB-month of changes)", 0.20, 0.20, "WAL retention for PITR"),
    ("Public network egress (per GB over 100 GB)", 0.10, 0.10, "First 100 GB included on paid"),
    ("Extra branch (per branch-month)", 1.50, 1.50, "Beyond 10 included (Launch) / 25 (Scale)"),
]
neon_start = r + 1
for label, lp, sp, src in neon:
    r += 1
    cel(ws, r, 1, label)
    cel(ws, r, 2, lp, bold, inp_fill, money3)
    cel(ws, r, 3, sp, bold, inp_fill, money3)
    cel(ws, r, 5, src, note_font)
# Reference cells for formulas
UP_COMPUTE_L = f"'Unit Prices'!B{neon_start}"      # 0.106
UP_COMPUTE_S = f"'Unit Prices'!C{neon_start}"      # 0.222
UP_STORAGE   = f"'Unit Prices'!B{neon_start+1}"    # 0.35
UP_RESTORE   = f"'Unit Prices'!B{neon_start+2}"    # 0.20
UP_EGRESS    = f"'Unit Prices'!B{neon_start+3}"    # 0.10

# Render rates
r += 2; sec(ws, r, MC, "RENDER RATES (render.com — current published tiers)")
r += 1
for c, h in enumerate(["Instance Type", "$/month", "CPU", "RAM (GB)", "Notes"], 1):
    cel(ws, r, c, h)
hdr(ws, r, MC)

render = [
    ("Free", 0, 0.1, 0.5, "Spins down 15min idle — dev only"),
    ("Starter", 7, 0.5, 0.5, "Background workers, light tasks"),
    ("Standard", 25, 1, 2, "1 CPU / 2 GB — good baseline"),
    ("Pro", 85, 2, 4, "2 CPU / 4 GB — medium production"),
    ("Pro Plus", 175, 4, 8, "4 CPU / 8 GB — high concurrency"),
    ("Pro Max", 225, 4, 16, "4 CPU / 16 GB — memory heavy"),
    ("Pro Ultra", 450, 8, 32, "8 CPU / 32 GB — enterprise"),
]
render_start = r + 1
for tier, cost, cpu, ram, note in render:
    r += 1
    cel(ws, r, 1, tier)
    cel(ws, r, 2, cost, bold, inp_fill, '$#,##0.00')
    cel(ws, r, 3, cpu, norm, None, dec1)
    cel(ws, r, 4, ram, norm, None, dec1)
    cel(ws, r, 5, note, note_font)

UP_R_STARTER  = f"'Unit Prices'!B{render_start+1}"  # $7
UP_R_STANDARD = f"'Unit Prices'!B{render_start+2}"  # $25
UP_R_PRO      = f"'Unit Prices'!B{render_start+3}"  # $85

# CU sizing reference
r += 2; sec(ws, r, MC, "NEON CU SIZING REFERENCE")
r += 1
for c, h in enumerate(["CU Size", "RAM", "Typical Use", "", ""], 1):
    cel(ws, r, c, h)
hdr(ws, r, MC)
cu_ref = [
    (0.25, "1 GB", "Dev/staging, very light queries"),
    (0.5, "2 GB", "Light production, small working set"),
    (1, "4 GB", "Medium production — 5 tenants baseline"),
    (2, "8 GB", "Active multi-tenant, complex joins"),
    (4, "16 GB", "Heavy reporting + OLTP concurrent"),
]
for cu, ram, use in cu_ref:
    r += 1
    cel(ws, r, 1, cu, norm, None, dec1 if cu < 1 else int_f)
    cel(ws, r, 2, ram)
    cel(ws, r, 3, use)


# ================================================================
# SHEET 2: DATA VOLUME
# ================================================================
ws2 = wb.create_sheet("Data Volume")
widths(ws2, [35, 16, 16, 16, 16, 30])
MC2 = 6

ws2.merge_cells('A1:F1')
cel(ws2, 1, 1, "DATA VOLUME — Records × Row Size = Storage", Font(name="Calibri", bold=True, size=14, color="2F5496"))
cel(ws2, 2, 1, "Estimates per company per year. Based on our 275 Prisma models (schema analysis).", note_font)

# Company count input
r = 4
cel(ws2, r, 1, "Number of companies", bold)
cel(ws2, r, 2, 5, bold, inp_fill, int_f)
COMPANIES = "B4"

r += 1
cel(ws2, r, 1, "Projection period", bold)
cel(ws2, r, 2, 1, bold, inp_fill, int_f)
cel(ws2, r, 3, "year(s)")
YEARS = "B5"

# Module breakdown
r += 2; sec(ws2, r, MC2, "RENT-A-CAR MODULE (64 Prisma models — bookings, contracts, vehicles, customers, dispatch)")
r += 1
for c, h in enumerate(["Table / Data Type", "Records/Co/Yr", "Avg Row (KB)", "GB/Co/Yr", "GB Total", "Basis"], 1):
    cel(ws2, r, c, h)
hdr(ws2, r, MC2)

rac_data = [
    ("vehicles", 200, 2.5, "~71 fields incl JSON metadata"),
    ("bookings", 6000, 1.8, "~60 fields, 500/mo × 12"),
    ("rentalContracts", 4800, 2.8, "~78 fields, 400/mo × 12"),
    ("customers", 2000, 2.0, "~59 fields, growing base"),
    ("deliveries + dispatch", 4800, 1.0, "Delivery slots, tracking logs"),
    ("deposits + charges", 3000, 0.5, "Operational charges per contract"),
    ("driver profiles + KYC docs", 500, 1.5, "Driver + document metadata"),
    ("vehicle_movements", 12000, 1.8, "~58 fields, plate/status changes"),
    ("booking_chat + notes", 8000, 0.8, "Messages, tags, internal notes"),
    ("Indexes overhead (this module)", None, None, "~40% of data size"),
]

rac_start = r + 1
for i, (table, recs, row_kb, basis) in enumerate(rac_data):
    r += 1
    cel(ws2, r, 1, table)
    if recs is not None:
        cel(ws2, r, 2, recs, norm, inp_fill, int_f)
        cel(ws2, r, 3, row_kb, norm, inp_fill, dec1)
        # GB/Co/Yr = records × row_kb / 1024 / 1024 (KB to GB)
        cel(ws2, r, 4, f"=B{r}*C{r}/1024/1024*{YEARS}", norm, calc_fill, dec3)
        # GB Total = GB/Co × companies
        cel(ws2, r, 5, f"=D{r}*{COMPANIES}", norm, calc_fill, dec3)
    else:
        # Index overhead row
        cel(ws2, r, 4, f"=SUM(D{rac_start}:D{r-1})*0.4", norm, calc_fill, dec3)
        cel(ws2, r, 5, f"=D{r}*{COMPANIES}", norm, calc_fill, dec3)
    cel(ws2, r, 6, basis, note_font)
rac_end = r

r += 1
cel(ws2, r, 1, "RENT-A-CAR SUBTOTAL", bold)
cel(ws2, r, 4, f"=SUM(D{rac_start}:D{rac_end})", bold, tot_fill, dec3)
cel(ws2, r, 5, f"=SUM(E{rac_start}:E{rac_end})", bold, tot_fill, dec3)
rac_total_per = f"D{r}"
rac_total_all = f"E{r}"

# Finance
r += 2; sec(ws2, r, MC2, "FINANCE MODULE (58 Prisma models — invoices, payments, vouchers, GL, assets)")
r += 1
for c, h in enumerate(["Table / Data Type", "Records/Co/Yr", "Avg Row (KB)", "GB/Co/Yr", "GB Total", "Basis"], 1):
    cel(ws2, r, c, h)
hdr(ws2, r, MC2)

fin_data = [
    ("invoices", 6000, 1.5, "~500/mo, linked to contracts"),
    ("payments", 6000, 1.2, "Payment records per invoice"),
    ("accounting_entries", 15000, 1.0, "~35 fields, journal entries"),
    ("receipt_vouchers", 3000, 1.5, "~45 fields"),
    ("payment_vouchers", 2000, 1.4, "~43 fields"),
    ("credit_notes", 500, 1.2, "~36 fields"),
    ("fixed_assets", 100, 2.0, "~56 fields, depreciation records"),
    ("transactions + GL", 20000, 0.8, "High-volume ledger lines"),
    ("Indexes overhead (this module)", None, None, "~40% of data size"),
]

fin_start = r + 1
for table, recs, row_kb, basis in fin_data:
    r += 1
    cel(ws2, r, 1, table)
    if recs is not None:
        cel(ws2, r, 2, recs, norm, inp_fill, int_f)
        cel(ws2, r, 3, row_kb, norm, inp_fill, dec1)
        cel(ws2, r, 4, f"=B{r}*C{r}/1024/1024*{YEARS}", norm, calc_fill, dec3)
        cel(ws2, r, 5, f"=D{r}*{COMPANIES}", norm, calc_fill, dec3)
    else:
        cel(ws2, r, 4, f"=SUM(D{fin_start}:D{r-1})*0.4", norm, calc_fill, dec3)
        cel(ws2, r, 5, f"=D{r}*{COMPANIES}", norm, calc_fill, dec3)
    cel(ws2, r, 6, basis, note_font)
fin_end = r

r += 1
cel(ws2, r, 1, "FINANCE SUBTOTAL", bold)
cel(ws2, r, 4, f"=SUM(D{fin_start}:D{fin_end})", bold, tot_fill, dec3)
cel(ws2, r, 5, f"=SUM(E{fin_start}:E{fin_end})", bold, tot_fill, dec3)
fin_total_per = f"D{r}"
fin_total_all = f"E{r}"

# TARS
r += 2; sec(ws2, r, MC2, "TARS MODULE (2 models — traffic fines + Salik toll charges)")
r += 1
for c, h in enumerate(["Table / Data Type", "Records/Co/Yr", "Avg Row (KB)", "GB/Co/Yr", "GB Total", "Basis"], 1):
    cel(ws2, r, c, h)
hdr(ws2, r, MC2)

tars_data = [
    ("tars_fines", 2400, 1.0, "~19 fields, 200/mo × 12"),
    ("tars_salik", 6000, 0.7, "~16 fields, toll charges daily"),
    ("Indexes overhead", None, None, "~40%"),
]
tars_start = r + 1
for table, recs, row_kb, basis in tars_data:
    r += 1
    cel(ws2, r, 1, table)
    if recs is not None:
        cel(ws2, r, 2, recs, norm, inp_fill, int_f)
        cel(ws2, r, 3, row_kb, norm, inp_fill, dec1)
        cel(ws2, r, 4, f"=B{r}*C{r}/1024/1024*{YEARS}", norm, calc_fill, dec3)
        cel(ws2, r, 5, f"=D{r}*{COMPANIES}", norm, calc_fill, dec3)
    else:
        cel(ws2, r, 4, f"=SUM(D{tars_start}:D{r-1})*0.4", norm, calc_fill, dec3)
        cel(ws2, r, 5, f"=D{r}*{COMPANIES}", norm, calc_fill, dec3)
    cel(ws2, r, 6, basis, note_font)
tars_end = r
r += 1
cel(ws2, r, 1, "TARS SUBTOTAL", bold)
cel(ws2, r, 4, f"=SUM(D{tars_start}:D{tars_end})", bold, tot_fill, dec3)
cel(ws2, r, 5, f"=SUM(E{tars_start}:E{tars_end})", bold, tot_fill, dec3)
tars_total_per = f"D{r}"
tars_total_all = f"E{r}"

# HR
r += 2; sec(ws2, r, MC2, "HR MODULE (2 models — payroll, salary records)")
r += 1
for c, h in enumerate(["Table / Data Type", "Records/Co/Yr", "Avg Row (KB)", "GB/Co/Yr", "GB Total", "Basis"], 1):
    cel(ws2, r, c, h)
hdr(ws2, r, MC2)

hr_data = [
    ("employee_salary_records", 600, 1.2, "~29 fields, 50 emp × 12 months"),
    ("payroll_cycles", 12, 1.0, "~27 fields, monthly cycle"),
    ("attendance + leave (est.)", 15000, 0.3, "Daily records × 50 employees"),
    ("Indexes overhead", None, None, "~40%"),
]
hr_start = r + 1
for table, recs, row_kb, basis in hr_data:
    r += 1
    cel(ws2, r, 1, table)
    if recs is not None:
        cel(ws2, r, 2, recs, norm, inp_fill, int_f)
        cel(ws2, r, 3, row_kb, norm, inp_fill, dec1)
        cel(ws2, r, 4, f"=B{r}*C{r}/1024/1024*{YEARS}", norm, calc_fill, dec3)
        cel(ws2, r, 5, f"=D{r}*{COMPANIES}", norm, calc_fill, dec3)
    else:
        cel(ws2, r, 4, f"=SUM(D{hr_start}:D{r-1})*0.4", norm, calc_fill, dec3)
        cel(ws2, r, 5, f"=D{r}*{COMPANIES}", norm, calc_fill, dec3)
    cel(ws2, r, 6, basis, note_font)
hr_end = r
r += 1
cel(ws2, r, 1, "HR SUBTOTAL", bold)
cel(ws2, r, 4, f"=SUM(D{hr_start}:D{hr_end})", bold, tot_fill, dec3)
cel(ws2, r, 5, f"=SUM(E{hr_start}:E{hr_end})", bold, tot_fill, dec3)
hr_total_per = f"D{r}"
hr_total_all = f"E{r}"

# Service Center + Recovery + Admin + Other
r += 2; sec(ws2, r, MC2, "SERVICE CENTER + RECOVERY + ADMIN + OTHER (145 models)")
r += 1
for c, h in enumerate(["Table / Data Type", "Records/Co/Yr", "Avg Row (KB)", "GB/Co/Yr", "GB Total", "Basis"], 1):
    cel(ws2, r, c, h)
hdr(ws2, r, MC2)

other_data = [
    ("service_booking + work_tasks", 1200, 1.5, "~32 fields, 100/mo"),
    ("spare_parts inventory", 500, 0.8, "~19 fields, catalog"),
    ("recovery_jobs", 600, 2.5, "~71 fields, 50/mo"),
    ("users + roles + permissions", 200, 3.0, "~89 fields on users"),
    ("audit_logs", 50000, 0.5, "High volume, every action logged"),
    ("notifications", 20000, 0.3, "System + email + SMS stubs"),
    ("sessions + tokens", 10000, 0.4, "Auth sessions, JWT records"),
    ("blog + support_tickets", 2000, 1.0, "~23 fields + 12 indexes on tickets"),
    ("Indexes overhead (all above)", None, None, "~40%"),
]
oth_start = r + 1
for table, recs, row_kb, basis in other_data:
    r += 1
    cel(ws2, r, 1, table)
    if recs is not None:
        cel(ws2, r, 2, recs, norm, inp_fill, int_f)
        cel(ws2, r, 3, row_kb, norm, inp_fill, dec1)
        cel(ws2, r, 4, f"=B{r}*C{r}/1024/1024*{YEARS}", norm, calc_fill, dec3)
        cel(ws2, r, 5, f"=D{r}*{COMPANIES}", norm, calc_fill, dec3)
    else:
        cel(ws2, r, 4, f"=SUM(D{oth_start}:D{r-1})*0.4", norm, calc_fill, dec3)
        cel(ws2, r, 5, f"=D{r}*{COMPANIES}", norm, calc_fill, dec3)
    cel(ws2, r, 6, basis, note_font)
oth_end = r
r += 1
cel(ws2, r, 1, "OTHER SUBTOTAL", bold)
cel(ws2, r, 4, f"=SUM(D{oth_start}:D{oth_end})", bold, tot_fill, dec3)
cel(ws2, r, 5, f"=SUM(E{oth_start}:E{oth_end})", bold, tot_fill, dec3)
oth_total_per = f"D{r}"
oth_total_all = f"E{r}"

# Grand total storage
r += 2
cel(ws2, r, 1, "GRAND TOTAL STORAGE", Font(name="Calibri", bold=True, size=12, color="2F5496"))
cel(ws2, r, 4, f"={rac_total_per}+{fin_total_per}+{tars_total_per}+{hr_total_per}+{oth_total_per}",
    Font(name="Calibri", bold=True, size=12), tot_fill, dec1)
cel(ws2, r, 5, f"={rac_total_all}+{fin_total_all}+{tars_total_all}+{hr_total_all}+{oth_total_all}",
    Font(name="Calibri", bold=True, size=12), tot_fill, dec1)
cel(ws2, r, 6, "GB", bold)
TOTAL_STORAGE_PER = f"'Data Volume'!D{r}"
TOTAL_STORAGE_ALL = f"'Data Volume'!E{r}"
COMPANIES_REF = f"'Data Volume'!{COMPANIES}"


# ================================================================
# SHEET 3: COMPUTE USAGE
# ================================================================
ws3 = wb.create_sheet("Compute Usage")
widths(ws3, [38, 16, 16, 16, 16, 30])
MC3 = 6

ws3.merge_cells('A1:F1')
cel(ws3, 1, 1, "COMPUTE USAGE — Hours × CU Size × Rate = Cost", Font(name="Calibri", bold=True, size=14, color="2F5496"))
cel(ws3, 2, 1, "Neon bills per CU-hour. More queries or heavier queries = higher CU = higher cost.", note_font)

# Time inputs
r = 4; sec(ws3, r, MC3, "TIME INPUTS")
r += 1
cel(ws3, r, 1, "Business hours per day", bold); cel(ws3, r, 2, 14, bold, inp_fill, int_f)
cel(ws3, r, 6, "UAE: ~8am to 10pm", note_font)
BIZ_H = f"B{r}"
r += 1
cel(ws3, r, 1, "Off-peak hours per day", bold); cel(ws3, r, 2, 10, bold, inp_fill, int_f)
cel(ws3, r, 6, "10pm to 8am", note_font)
OFF_H = f"B{r}"
r += 1
cel(ws3, r, 1, "Days per month", bold); cel(ws3, r, 2, 30, bold, inp_fill, int_f)
DAYS = f"B{r}"
r += 1
cel(ws3, r, 1, "Business hours/month", bold)
cel(ws3, r, 2, f"={BIZ_H}*{DAYS}", bold, calc_fill, int_f)
BIZ_MO = f"B{r}"
r += 1
cel(ws3, r, 1, "Off-peak hours/month", bold)
cel(ws3, r, 2, f"={OFF_H}*{DAYS}", bold, calc_fill, int_f)
OFF_MO = f"B{r}"

# Option A
r += 2; sec(ws3, r, MC3, "OPTION A: STARTER (Neon Launch)")
r += 1
for c, h in enumerate(["Period", "CU Size", "Hours", "CU-Hours", "× Rate", "Cost"], 1):
    cel(ws3, r, c, h)
hdr(ws3, r, MC3)

r += 1
cel(ws3, r, 1, "Business hours (avg load)")
cel(ws3, r, 2, 1, norm, inp_fill, dec1); a_biz_cu = f"B{r}"
cel(ws3, r, 3, f"={BIZ_MO}", norm, calc_fill, int_f)
cel(ws3, r, 4, f"=B{r}*C{r}", norm, calc_fill, int_f)
cel(ws3, r, 5, f"={UP_COMPUTE_L}", norm, None, money3)
cel(ws3, r, 6, f"=D{r}*E{r}", norm, calc_fill, money)
a_biz = f"F{r}"

r += 1
cel(ws3, r, 1, "Off-peak (minimal, before scale-to-zero)")
cel(ws3, r, 2, 0.25, norm, inp_fill, dec1); a_off_cu = f"B{r}"
cel(ws3, r, 3, f"={OFF_MO}", norm, calc_fill, int_f)
cel(ws3, r, 4, f"=B{r}*C{r}", norm, calc_fill, int_f)
cel(ws3, r, 5, f"={UP_COMPUTE_L}", norm, None, money3)
cel(ws3, r, 6, f"=D{r}*E{r}", norm, calc_fill, money)
a_off = f"F{r}"

r += 1
cel(ws3, r, 1, "OPTION A — COMPUTE TOTAL", bold)
cel(ws3, r, 4, f"=D{r-2}+D{r-1}", bold, None, int_f)
cel(ws3, r, 6, f"={a_biz}+{a_off}", bold, tot_fill, money)
A_COMPUTE = f"F{r}"
cel(ws3, r, 5, "CU-hrs total", note_font)

# Option B
r += 2; sec(ws3, r, MC3, "OPTION B: GROWTH (Neon Launch, higher CU)")
r += 1
for c, h in enumerate(["Period", "CU Size", "Hours", "CU-Hours", "× Rate", "Cost"], 1):
    cel(ws3, r, c, h)
hdr(ws3, r, MC3)

r += 1
cel(ws3, r, 1, "Business hours (heavier queries)")
cel(ws3, r, 2, 2, norm, inp_fill, dec1)
cel(ws3, r, 3, f"={BIZ_MO}", norm, calc_fill, int_f)
cel(ws3, r, 4, f"=B{r}*C{r}", norm, calc_fill, int_f)
cel(ws3, r, 5, f"={UP_COMPUTE_L}", norm, None, money3)
cel(ws3, r, 6, f"=D{r}*E{r}", norm, calc_fill, money)
b_biz = f"F{r}"

r += 1
cel(ws3, r, 1, "Off-peak")
cel(ws3, r, 2, 0.5, norm, inp_fill, dec1)
cel(ws3, r, 3, f"={OFF_MO}", norm, calc_fill, int_f)
cel(ws3, r, 4, f"=B{r}*C{r}", norm, calc_fill, int_f)
cel(ws3, r, 5, f"={UP_COMPUTE_L}", norm, None, money3)
cel(ws3, r, 6, f"=D{r}*E{r}", norm, calc_fill, money)
b_off = f"F{r}"

r += 1
cel(ws3, r, 1, "OPTION B — COMPUTE TOTAL", bold)
cel(ws3, r, 4, f"=D{r-2}+D{r-1}", bold, None, int_f)
cel(ws3, r, 6, f"={b_biz}+{b_off}", bold, tot_fill, money)
B_COMPUTE = f"F{r}"

# Option C
r += 2; sec(ws3, r, MC3, "OPTION C: ENTERPRISE (Neon Scale, always-on + replica)")
r += 1
for c, h in enumerate(["Period", "CU Size", "Hours", "CU-Hours", "× Rate", "Cost"], 1):
    cel(ws3, r, c, h)
hdr(ws3, r, MC3)

r += 1
cel(ws3, r, 1, "Primary compute (always-on)")
cel(ws3, r, 2, 2, norm, inp_fill, dec1)
cel(ws3, r, 3, 730, norm, inp_fill, int_f)
cel(ws3, r, 4, f"=B{r}*C{r}", norm, calc_fill, int_f)
cel(ws3, r, 5, f"={UP_COMPUTE_S}", norm, None, money3)
cel(ws3, r, 6, f"=D{r}*E{r}", norm, calc_fill, money)
c_pri = f"F{r}"

r += 1
cel(ws3, r, 1, "Read replica (reporting)")
cel(ws3, r, 2, 0.5, norm, inp_fill, dec1)
cel(ws3, r, 3, 730, norm, inp_fill, int_f)
cel(ws3, r, 4, f"=B{r}*C{r}", norm, calc_fill, int_f)
cel(ws3, r, 5, f"={UP_COMPUTE_S}", norm, None, money3)
cel(ws3, r, 6, f"=D{r}*E{r}", norm, calc_fill, money)
c_rep = f"F{r}"

r += 1
cel(ws3, r, 1, "OPTION C — COMPUTE TOTAL", bold)
cel(ws3, r, 4, f"=D{r-2}+D{r-1}", bold, None, int_f)
cel(ws3, r, 6, f"={c_pri}+{c_rep}", bold, tot_fill, money)
C_COMPUTE = f"F{r}"


# ================================================================
# SHEET 4: COST ROLLUP (Qty × Price for everything)
# ================================================================
ws4 = wb.create_sheet("Cost Rollup")
widths(ws4, [38, 16, 16, 16, 16, 16])
MC4 = 6

ws4.merge_cells('A1:F1')
cel(ws4, 1, 1, "COST ROLLUP — Every line = Quantity × Unit Price", Font(name="Calibri", bold=True, size=14, color="2F5496"))
cel(ws4, 2, 1, "No lump sums. Change any input on prior sheets and costs recalculate.", note_font)

for opt_label, opt_letter, compute_ref, neon_plan, render_backend, render_backend_n, render_worker, render_worker_n, team_seats, restore_gb in [
    ("OPTION A: STARTER", "A", f"'Compute Usage'!{A_COMPUTE}", "Launch", UP_R_STANDARD, 1, UP_R_STARTER, 1, 0, 5),
    ("OPTION B: GROWTH", "B", f"'Compute Usage'!{B_COMPUTE}", "Launch", UP_R_PRO, 1, UP_R_STANDARD, 1, 1, 8),
    ("OPTION C: ENTERPRISE", "C", f"'Compute Usage'!{C_COMPUTE}", "Scale", UP_R_PRO, 2, UP_R_STANDARD, 1, 3, 25),
]:
    r = ws4.max_row + 2 if ws4.max_row > 2 else 4
    sec(ws4, r, MC4, opt_label)
    r += 1
    for c, h in enumerate(["Line Item", "Quantity", "Unit", "Unit Price", "Subtotal", "Notes"], 1):
        cel(ws4, r, c, h)
    hdr(ws4, r, MC4)

    start_r = r + 1

    # Neon Compute
    r += 1
    cel(ws4, r, 1, f"Neon Compute ({neon_plan})")
    cel(ws4, r, 2, f"={compute_ref.replace('F','D').replace('Cost Rollup','Compute Usage')}", norm, calc_fill, int_f)
    # Actually, let's just reference the total CU-hours and cost directly
    cel(ws4, r, 2, 1, norm, None, int_f)
    cel(ws4, r, 3, "lump (see Compute Usage)")
    cel(ws4, r, 5, f"={compute_ref}", norm, calc_fill, money)
    cel(ws4, r, 6, "CU-hours × rate — see Compute Usage sheet", note_font)
    neon_compute_row = f"E{r}"

    # Neon Storage
    r += 1
    cel(ws4, r, 1, "Neon Storage")
    cel(ws4, r, 2, f"={TOTAL_STORAGE_ALL}", norm, calc_fill, dec1)
    cel(ws4, r, 3, "GB-month")
    cel(ws4, r, 4, f"={UP_STORAGE}", norm, None, money3)
    cel(ws4, r, 5, f"=B{r}*D{r}", norm, calc_fill, money)
    cel(ws4, r, 6, "From Data Volume sheet", note_font)
    neon_storage_row = f"E{r}"

    # Neon Restore
    r += 1
    cel(ws4, r, 1, "Neon Instant Restore")
    cel(ws4, r, 2, restore_gb, norm, inp_fill, dec1)
    cel(ws4, r, 3, "GB-month (changes)")
    cel(ws4, r, 4, f"={UP_RESTORE}", norm, None, money3)
    cel(ws4, r, 5, f"=B{r}*D{r}", norm, calc_fill, money)
    cel(ws4, r, 6, "WAL retention for point-in-time recovery", note_font)
    neon_restore_row = f"E{r}"

    # Neon egress
    r += 1
    cel(ws4, r, 1, "Neon Network Egress (over 100 GB)")
    cel(ws4, r, 2, 0, norm, inp_fill, dec1)
    cel(ws4, r, 3, "GB over included")
    cel(ws4, r, 4, f"={UP_EGRESS}", norm, None, money3)
    cel(ws4, r, 5, f"=B{r}*D{r}", norm, calc_fill, money)
    cel(ws4, r, 6, "100 GB included free on paid plans", note_font)
    neon_egress_row = f"E{r}"

    # Neon subtotal
    r += 1
    cel(ws4, r, 1, "NEON SUBTOTAL", bold)
    cel(ws4, r, 5, f"={neon_compute_row}+{neon_storage_row}+{neon_restore_row}+{neon_egress_row}", bold, tot_fill, money)
    neon_sub = f"E{r}"

    # Render Backend
    r += 1
    cel(ws4, r, 1, f"Render Backend API")
    cel(ws4, r, 2, render_backend_n, norm, inp_fill, int_f)
    cel(ws4, r, 3, "instance(s)")
    cel(ws4, r, 4, f"={render_backend}", norm, None, '$#,##0.00')
    cel(ws4, r, 5, f"=B{r}*D{r}", norm, calc_fill, money)
    render_be_row = f"E{r}"

    # Render Frontend
    r += 1
    cel(ws4, r, 1, "Render Frontend (Static Site)")
    cel(ws4, r, 2, 1, norm, None, int_f)
    cel(ws4, r, 3, "site")
    cel(ws4, r, 4, 0, norm, None, '$#,##0.00')
    cel(ws4, r, 5, f"=B{r}*D{r}", norm, calc_fill, money)
    cel(ws4, r, 6, "Free tier", note_font)
    render_fe_row = f"E{r}"

    # Render Worker
    r += 1
    cel(ws4, r, 1, "Render Background Worker")
    cel(ws4, r, 2, render_worker_n, norm, inp_fill, int_f)
    cel(ws4, r, 3, "instance(s)")
    cel(ws4, r, 4, f"={render_worker}", norm, None, '$#,##0.00')
    cel(ws4, r, 5, f"=B{r}*D{r}", norm, calc_fill, money)
    render_wk_row = f"E{r}"

    # Team seats
    r += 1
    cel(ws4, r, 1, "Render Team Workspace")
    cel(ws4, r, 2, team_seats, norm, inp_fill, int_f)
    cel(ws4, r, 3, "seat(s)")
    cel(ws4, r, 4, 19, norm, inp_fill, '$#,##0.00')
    cel(ws4, r, 5, f"=B{r}*D{r}", norm, calc_fill, money)
    cel(ws4, r, 6, "$19/seat/mo (0 = Individual plan)", note_font)
    render_tm_row = f"E{r}"

    # Render subtotal
    r += 1
    cel(ws4, r, 1, "RENDER SUBTOTAL", bold)
    cel(ws4, r, 5, f"={render_be_row}+{render_fe_row}+{render_wk_row}+{render_tm_row}", bold, tot_fill, money)
    render_sub = f"E{r}"

    # Grand total for this option
    r += 1
    for c in range(1, MC4+1):
        ws4.cell(r, c).fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        ws4.cell(r, c).border = bdr
    cel(ws4, r, 1, f"{opt_label} — TOTAL / MONTH", Font(name="Calibri", bold=True, size=12, color="FFFFFF"))
    cel(ws4, r, 5, f"={neon_sub}+{render_sub}", Font(name="Calibri", bold=True, size=12, color="FFFFFF"), None, '$#,##0.00')

    r += 1
    cel(ws4, r, 1, "Total / Year", bold)
    cel(ws4, r, 5, f"=E{r-1}*12", bold, tot_fill, '$#,##0.00')
    r += 1
    cel(ws4, r, 1, "Per Company / Month", bold)
    cel(ws4, r, 5, f"=E{r-2}/{COMPANIES_REF}", bold, tot_fill, '$#,##0.00')


# ================================================================
# SHEET 5: PER CLIENT BREAKDOWN (Company A–E)
# ================================================================
ws5 = wb.create_sheet("Per Client")
widths(ws5, [32, 16, 16, 16, 16, 16, 16, 16])
MC5 = 8

ws5.merge_cells('A1:H1')
cel(ws5, 1, 1, "PER CLIENT COST ESTIMATE — Companies A through E",
    Font(name="Calibri", bold=True, size=14, color="2F5496"))
cel(ws5, 2, 1,
    "Each company has its own fleet size, user count, and data profile. Costs are proportional to usage.",
    note_font)

# --- Client profiles ---
r = 4; sec(ws5, r, MC5, "CLIENT PROFILES (edit yellow cells per company)")
r += 1
for c, h in enumerate(["Metric", "Company A", "Company B", "Company C", "Company D", "Company E", "TOTAL", "Notes"], 1):
    cel(ws5, r, c, h)
hdr(ws5, r, MC5)

# Row: Company names
r += 1
cel(ws5, r, 1, "Company Name", bold)
for ci, name in enumerate(["Al Raha Fleet", "Desert Star Rental", "Gulf Line Cars", "Oasis Motors", "Palm Drive LLC"], 2):
    cel(ws5, r, ci, name, bold, inp_fill)
cel(ws5, r, 8, "Example names — edit to match actual clients", note_font)
name_row = r

# Fleet size
r += 1
cel(ws5, r, 1, "Fleet size (vehicles)", bold)
for ci, val in enumerate([120, 250, 80, 300, 150], 2):
    cel(ws5, r, ci, val, norm, inp_fill, int_f)
cel(ws5, r, 7, f"=SUM(B{r}:F{r})", bold, calc_fill, int_f)
fleet_row = r

# Concurrent users
r += 1
cel(ws5, r, 1, "Concurrent users (peak)", bold)
for ci, val in enumerate([10, 25, 8, 30, 15], 2):
    cel(ws5, r, ci, val, norm, inp_fill, int_f)
cel(ws5, r, 7, f"=SUM(B{r}:F{r})", bold, calc_fill, int_f)
users_row = r

# Contracts per month (≈ fleet × 60-75% utilization, avg 2-week rental)
r += 1
cel(ws5, r, 1, "Contracts / month", bold)
for ci, val in enumerate([70, 160, 45, 200, 90], 2):
    cel(ws5, r, ci, val, norm, inp_fill, int_f)
cel(ws5, r, 7, f"=SUM(B{r}:F{r})", bold, calc_fill, int_f)
cel(ws5, r, 8, "~50-70% of fleet, mix of short/long-term", note_font)
contracts_row = r

# Invoices per month (≈ contracts + extensions/charges)
r += 1
cel(ws5, r, 1, "Invoices / month", bold)
for ci, val in enumerate([85, 190, 55, 240, 110], 2):
    cel(ws5, r, ci, val, norm, inp_fill, int_f)
cel(ws5, r, 7, f"=SUM(B{r}:F{r})", bold, calc_fill, int_f)
cel(ws5, r, 8, "Contracts + extensions + additional charges", note_font)
invoices_row = r

# TARS fines per month
r += 1
cel(ws5, r, 1, "TARS fines / month", bold)
for ci, val in enumerate([80, 200, 50, 250, 120], 2):
    cel(ws5, r, ci, val, norm, inp_fill, int_f)
cel(ws5, r, 7, f"=SUM(B{r}:F{r})", bold, calc_fill, int_f)
fines_row = r

# Employees
r += 1
cel(ws5, r, 1, "Employees (HR)", bold)
for ci, val in enumerate([25, 60, 15, 80, 40], 2):
    cel(ws5, r, ci, val, norm, inp_fill, int_f)
cel(ws5, r, 7, f"=SUM(B{r}:F{r})", bold, calc_fill, int_f)
emp_row = r

# Customers
r += 1
cel(ws5, r, 1, "Customers in DB", bold)
for ci, val in enumerate([800, 2000, 500, 3000, 1200], 2):
    cel(ws5, r, ci, val, norm, inp_fill, int_f)
cel(ws5, r, 7, f"=SUM(B{r}:F{r})", bold, calc_fill, int_f)
cust_row = r

# --- Storage per client ---
r += 2; sec(ws5, r, MC5, "ESTIMATED STORAGE PER CLIENT (GB) — records × avg row size + 40% index overhead")
r += 1
for c, h in enumerate(["Data Category", "Company A", "Company B", "Company C", "Company D", "Company E", "TOTAL", "Calculation"], 1):
    cel(ws5, r, c, h)
hdr(ws5, r, MC5)

# Vehicles + fleet data: fleet × 2.5 KB + fleet × 12 movements/yr × 1.8 KB
r += 1
cel(ws5, r, 1, "Vehicles + movements")
for ci in range(2, 7):
    col_l = get_column_letter(ci)
    # (fleet × 2.5 KB) + (fleet × 12 × 1.8 KB) → KB → GB
    cel(ws5, r, ci, f"=({col_l}{fleet_row}*2.5 + {col_l}{fleet_row}*12*1.8)/1024/1024", norm, calc_fill, dec3)
cel(ws5, r, 7, f"=SUM(B{r}:F{r})", bold, calc_fill, dec3)
cel(ws5, r, 8, "fleet × 2.5KB + fleet × 12mo × 1.8KB movement", note_font)
stor_veh = r

# Contracts + bookings: contracts × 12 × (2.8 + 1.8) KB
r += 1
cel(ws5, r, 1, "Contracts + bookings")
for ci in range(2, 7):
    col_l = get_column_letter(ci)
    cel(ws5, r, ci, f"={col_l}{contracts_row}*12*(2.8+1.8)/1024/1024", norm, calc_fill, dec3)
cel(ws5, r, 7, f"=SUM(B{r}:F{r})", bold, calc_fill, dec3)
cel(ws5, r, 8, "contracts/mo × 12 × (2.8+1.8) KB per booking+contract", note_font)
stor_con = r

# Customers + KYC: customers × 2.0 KB + customers × 0.5 × 1.5 KB (docs)
r += 1
cel(ws5, r, 1, "Customers + KYC docs")
for ci in range(2, 7):
    col_l = get_column_letter(ci)
    cel(ws5, r, ci, f"=({col_l}{cust_row}*2.0 + {col_l}{cust_row}*0.5*1.5)/1024/1024", norm, calc_fill, dec3)
cel(ws5, r, 7, f"=SUM(B{r}:F{r})", bold, calc_fill, dec3)
cel(ws5, r, 8, "customers × 2KB + 50% have docs × 1.5KB", note_font)
stor_cust = r

# Invoices + payments + GL: invoices × 12 × (1.5 + 1.2 + 3 × 0.8) KB
r += 1
cel(ws5, r, 1, "Invoices + payments + GL")
for ci in range(2, 7):
    col_l = get_column_letter(ci)
    cel(ws5, r, ci, f"={col_l}{invoices_row}*12*(1.5+1.2+3*0.8)/1024/1024", norm, calc_fill, dec3)
cel(ws5, r, 7, f"=SUM(B{r}:F{r})", bold, calc_fill, dec3)
cel(ws5, r, 8, "inv/mo × 12 × (1.5 inv + 1.2 pay + 2.4 journal KB)", note_font)
stor_fin = r

# TARS: fines × 12 × 1.0 KB + fines × 2.5 × 12 × 0.7 KB (salik ≈ 2.5× fines)
r += 1
cel(ws5, r, 1, "TARS fines + Salik")
for ci in range(2, 7):
    col_l = get_column_letter(ci)
    cel(ws5, r, ci, f"=({col_l}{fines_row}*12*1.0 + {col_l}{fines_row}*2.5*12*0.7)/1024/1024", norm, calc_fill, dec3)
cel(ws5, r, 7, f"=SUM(B{r}:F{r})", bold, calc_fill, dec3)
cel(ws5, r, 8, "fines/mo × 12 × 1KB + salik (2.5× fines) × 0.7KB", note_font)
stor_tars = r

# HR: employees × 12 × 1.2 KB (salary) + employees × 300 × 0.3 KB (attendance)
r += 1
cel(ws5, r, 1, "HR + payroll + attendance")
for ci in range(2, 7):
    col_l = get_column_letter(ci)
    cel(ws5, r, ci, f"=({col_l}{emp_row}*12*1.2 + {col_l}{emp_row}*300*0.3)/1024/1024", norm, calc_fill, dec3)
cel(ws5, r, 7, f"=SUM(B{r}:F{r})", bold, calc_fill, dec3)
cel(ws5, r, 8, "emp × 12 salary recs + emp × 300 workdays × 0.3KB attend", note_font)
stor_hr = r

# Service + Recovery + Admin + Audit (proportional to users)
r += 1
cel(ws5, r, 1, "Service + Recovery + Admin + Logs")
for ci in range(2, 7):
    col_l = get_column_letter(ci)
    # audit logs ≈ users × 2000 actions/yr × 0.5 KB + service/recovery/admin baseline ~0.3 GB
    cel(ws5, r, ci, f"=({col_l}{users_row}*2000*0.5)/1024/1024 + 0.3", norm, calc_fill, dec3)
cel(ws5, r, 7, f"=SUM(B{r}:F{r})", bold, calc_fill, dec3)
cel(ws5, r, 8, "users × 2000 actions × 0.5KB logs + 0.3GB baseline", note_font)
stor_other = r

# Index overhead
r += 1
cel(ws5, r, 1, "Index overhead (40%)")
for ci in range(2, 7):
    cel(ws5, r, ci, f"=SUM({get_column_letter(ci)}{stor_veh}:{get_column_letter(ci)}{stor_other})*0.4", norm, calc_fill, dec3)
cel(ws5, r, 7, f"=SUM(B{r}:F{r})", bold, calc_fill, dec3)
cel(ws5, r, 8, "628 indexes — ~40% overhead is typical for write-heavy ERP", note_font)
stor_idx = r

# Storage total per client
r += 1
cel(ws5, r, 1, "TOTAL STORAGE (GB)", Font(name="Calibri", bold=True, size=11, color="2F5496"))
for ci in range(2, 7):
    cel(ws5, r, ci, f"=SUM({get_column_letter(ci)}{stor_veh}:{get_column_letter(ci)}{stor_idx})",
        bold, tot_fill, dec1)
cel(ws5, r, 7, f"=SUM(B{r}:F{r})", Font(name="Calibri", bold=True, size=12), tot_fill, dec1)
storage_total_row = r

# --- Cost allocation per client ---
r += 2; sec(ws5, r, MC5, "MONTHLY COST PER CLIENT (Option A: Starter — Neon Launch + Render Standard)")
r += 1
for c, h in enumerate(["Cost Component", "Company A", "Company B", "Company C", "Company D", "Company E", "TOTAL", "How Calculated"], 1):
    cel(ws5, r, c, h)
hdr(ws5, r, MC5)

# Storage cost: per client GB × $0.35/GB
r += 1
cel(ws5, r, 1, "Neon Storage")
for ci in range(2, 7):
    col_l = get_column_letter(ci)
    cel(ws5, r, ci, f"={col_l}{storage_total_row}*{UP_STORAGE}", norm, calc_fill, money)
cel(ws5, r, 7, f"=SUM(B{r}:F{r})", bold, calc_fill, money)
cel(ws5, r, 8, "Client GB × $0.35/GB-month", note_font)
cost_storage_row = r

# Compute cost: proportional to user share (users drive query load)
r += 1
cel(ws5, r, 1, "Neon Compute (by user share)")
for ci in range(2, 7):
    col_l = get_column_letter(ci)
    # user share × total compute cost from Compute Usage sheet Option A
    cel(ws5, r, ci,
        f"=({col_l}{users_row}/{get_column_letter(7)}{users_row})*'Compute Usage'!{A_COMPUTE}",
        norm, calc_fill, money)
cel(ws5, r, 7, f"=SUM(B{r}:F{r})", bold, calc_fill, money)
cel(ws5, r, 8, "(client users ÷ total users) × total compute cost", note_font)
cost_compute_row = r

# Restore cost: proportional to storage
r += 1
cel(ws5, r, 1, "Neon Instant Restore")
for ci in range(2, 7):
    col_l = get_column_letter(ci)
    # proportional to storage share × 5 GB total estimated changes × $0.20
    cel(ws5, r, ci,
        f"=({col_l}{storage_total_row}/{get_column_letter(7)}{storage_total_row})*5*{UP_RESTORE}",
        norm, calc_fill, money)
cel(ws5, r, 7, f"=SUM(B{r}:F{r})", bold, calc_fill, money)
cel(ws5, r, 8, "(client storage ÷ total) × 5GB changes × $0.20", note_font)
cost_restore_row = r

# Render: shared infra — split evenly or by users
r += 1
cel(ws5, r, 1, "Render Backend (by user share)")
for ci in range(2, 7):
    col_l = get_column_letter(ci)
    cel(ws5, r, ci,
        f"=({col_l}{users_row}/{get_column_letter(7)}{users_row})*{UP_R_STANDARD}",
        norm, calc_fill, money)
cel(ws5, r, 7, f"=SUM(B{r}:F{r})", bold, calc_fill, money)
cel(ws5, r, 8, "(client users ÷ total) × $25 Standard instance", note_font)
cost_render_be_row = r

r += 1
cel(ws5, r, 1, "Render Worker (split evenly)")
for ci in range(2, 7):
    cel(ws5, r, ci, f"={UP_R_STARTER}/5", norm, calc_fill, money)
cel(ws5, r, 7, f"=SUM(B{r}:F{r})", bold, calc_fill, money)
cel(ws5, r, 8, "$7 Starter ÷ 5 companies", note_font)
cost_render_wk_row = r

r += 1
cel(ws5, r, 1, "Render Frontend (split evenly)")
for ci in range(2, 7):
    cel(ws5, r, ci, 0, norm, calc_fill, money)
cel(ws5, r, 7, 0, bold, calc_fill, money)
cel(ws5, r, 8, "Static site = free", note_font)
cost_render_fe_row = r

# TOTAL per client
r += 1
cel(ws5, r, 1, "TOTAL COST / MONTH", Font(name="Calibri", bold=True, size=12, color="2F5496"))
for ci in range(2, 7):
    col_l = get_column_letter(ci)
    cel(ws5, r, ci,
        f"={col_l}{cost_storage_row}+{col_l}{cost_compute_row}+{col_l}{cost_restore_row}+{col_l}{cost_render_be_row}+{col_l}{cost_render_wk_row}+{col_l}{cost_render_fe_row}",
        Font(name="Calibri", bold=True, size=12), tot_fill, '$#,##0.00')
cel(ws5, r, 7,
    f"=SUM(B{r}:F{r})",
    Font(name="Calibri", bold=True, size=12), tot_fill, '$#,##0.00')
monthly_total_row = r

# TOTAL per year
r += 1
cel(ws5, r, 1, "TOTAL COST / YEAR", bold)
for ci in range(2, 7):
    cel(ws5, r, ci, f"={get_column_letter(ci)}{monthly_total_row}*12", bold, tot_fill, '$#,##0.00')
cel(ws5, r, 7, f"=SUM(B{r}:F{r})", bold, tot_fill, '$#,##0.00')
yearly_total_row = r

# % share
r += 2; sec(ws5, r, MC5, "COST SHARE (%) — who's using what portion of the infrastructure")
r += 1
for c, h in enumerate(["Metric", "Company A", "Company B", "Company C", "Company D", "Company E", "TOTAL", ""], 1):
    cel(ws5, r, c, h)
hdr(ws5, r, MC5)

r += 1
cel(ws5, r, 1, "% of total cost")
for ci in range(2, 7):
    col_l = get_column_letter(ci)
    cel(ws5, r, ci, f"={col_l}{monthly_total_row}/G{monthly_total_row}", norm, calc_fill, '0.0%')
cel(ws5, r, 7, f"=SUM(B{r}:F{r})", bold, calc_fill, '0.0%')

r += 1
cel(ws5, r, 1, "% of storage")
for ci in range(2, 7):
    col_l = get_column_letter(ci)
    cel(ws5, r, ci, f"={col_l}{storage_total_row}/G{storage_total_row}", norm, calc_fill, '0.0%')
cel(ws5, r, 7, f"=SUM(B{r}:F{r})", bold, calc_fill, '0.0%')

r += 1
cel(ws5, r, 1, "% of compute (by users)")
for ci in range(2, 7):
    col_l = get_column_letter(ci)
    cel(ws5, r, ci, f"={col_l}{users_row}/G{users_row}", norm, calc_fill, '0.0%')
cel(ws5, r, 7, f"=SUM(B{r}:F{r})", bold, calc_fill, '0.0%')

# Tip
r += 2
ws5.merge_cells(f'A{r}:H{r}')
cel(ws5, r, 1,
    "💡 Edit the yellow profile cells above to match real client data. Storage, compute share, and costs all recalculate.",
    Font(name="Calibri", italic=True, size=11, color="2F5496"))


# ================================================================
# SAVE
# ================================================================
out = "/Users/miguelitodeguzman/Projects/tech-project/docs/Vesla_ERP_Cloud_Cost_Projection_v2.xlsx"
wb.save(out)
print(f"✅ Saved: {out}")
